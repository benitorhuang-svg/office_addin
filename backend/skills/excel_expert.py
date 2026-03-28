import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelExpertSkill:
    """
    Industrial Excel Hub: High-fidelity data manipulation and reporting.
    Powered by openpyxl for the Nexus Center platform.
    """
    def __init__(self, excel_path: str = None):
        self.wb = openpyxl.load_workbook(excel_path) if excel_path else openpyxl.Workbook()
        self.ws = self.wb.active

    def create_report(self, title: str, headers: list, data: list):
        # Apply High-Fidelity Professional Header
        self.ws['A1'] = title
        self.ws['A1'].font = Font(size=14, bold=True, name='Outfit')
        self.ws['A1'].alignment = Alignment(horizontal='center')
        
        # Style Headers (Nexus-Blue Fill)
        blue_fill = PatternFill(start_color="008CA1", end_color="008CA1", fill_type="solid")
        for col, h_text in enumerate(headers, 1):
            cell = self.ws.cell(row=2, column=col)
            cell.value = h_text
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = blue_fill
            
        # Insert Data
        for r_idx, row in enumerate(data, 3):
            for c_idx, val in enumerate(row, 1):
                self.ws.cell(row=r_idx, column=c_idx).value = val
                
    def save_workbook(self, output_path: str):
        self.wb.save(output_path)
        return {"status": "success", "file": output_path}

if __name__ == "__main__":
    try:
        # Standard ACP JSON Over-STDIN Pipe
        input_data = json.load(sys.stdin)
        action = input_data.get("action")
        
        expert = ExcelExpertSkill(input_data.get("path"))
        
        if action == "create_report":
            expert.create_report(input_data["title"], input_data["headers"], input_data["data"])
            res = expert.save_workbook(input_data["output"])
            print(json.dumps(res))
            
    except Exception as e:
        print(json.dumps({"error": str(e)}), file=sys.stderr)
        sys.exit(1)
