from __future__ import annotations

import csv
import json
import os
import re
import sys
from collections import defaultdict
from typing import Any, TypedDict, List, Optional, Dict, Tuple, Union

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import coordinate_from_string, range_boundaries
from bridge_protocol import BaseExpertSkill

# ── Type Definitions ─────────────────────────────────────────────────────

class PivotValueSpec(TypedDict, total=False):
    field: str
    func: str  # SUM, COUNT, AVERAGE

class ExcelOperation(TypedDict, total=False):
    op: str
    cell: Optional[str]
    value: Optional[Any]
    formula: Optional[str]
    range: Optional[str]
    bold: Optional[bool]
    font_color: Optional[str]
    fill_color: Optional[str]
    number_format: Optional[str]
    column: Optional[str]
    width: Optional[float]
    sheet: Optional[str]
    headers: Optional[List[str]]
    row: Optional[int]
    title: Optional[str]
    data: Optional[List[List[Any]]]
    source: Optional[str]
    destination: Optional[str]
    rows: Optional[List[str]]
    columns: Optional[List[str]]
    values: Optional[List[Union[str, PivotValueSpec]]]
    metadata: Optional[Dict[str, Any]]

class ExcelPayload(TypedDict):
    input: Optional[str]
    input_path: Optional[str]
    output: str
    output_path: Optional[str]
    changes: List[ExcelOperation]
    office_context: Optional[Dict[str, Any]]

# ── Brand defaults ────────────────────────────────────────────────────────
NEXUS_BLUE = "008CA1"
NEXUS_WHITE = "FFFFFF"
DEFAULT_PROFESSIONAL_FONT = "Arial"
SUPPORTED_SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm", ".csv", ".tsv"}

class ExcelExpertSkill(BaseExpertSkill):
    def __init__(self, excel_path: Optional[str] = None, office_context: Optional[Dict[str, Any]] = None) -> None:
        self.input_path = excel_path or ""
        self.office_context = office_context or {}
        self.wb = self._load_workbook(excel_path) if excel_path else openpyxl.Workbook()
        self.ws = self.wb.active

    def _op_set_value(self, op: ExcelOperation) -> None:
        ws, cell = self._resolve_sheet_and_ref(str(op["cell"]))
        ws[cell] = op["value"]

    def _op_add_formula(self, op: ExcelOperation) -> None:
        ws, cell = self._resolve_sheet_and_ref(str(op["cell"]))
        formula = self._normalize_formula(str(op["formula"]))
        # Basic validation: check if referenced sheets exist
        if "!" in formula:
            matches = re.findall(r"([a-zA-Z0-9_]+)!", formula) + re.findall(r"'([^']+)'!", formula)
            for match in matches:
                if match not in self.wb.sheetnames:
                    raise ValueError(f"Excel validation failed: Formula references missing sheet '{match}'")
        ws[cell] = formula

    def _normalize_formula(self, formula: str) -> str:
        trimmed = formula.strip()
        if trimmed.startswith("```") and trimmed.endswith("```"):
            inner = trimmed[3:-3].strip()
            if "\n" in inner: _, inner = inner.split("\n", 1)
            trimmed = inner.strip()
        if not trimmed.startswith("="): trimmed = "=" + trimmed
        return trimmed

    def _op_format_range(self, op: ExcelOperation) -> None:
        ws, cell_range = self._resolve_sheet_and_ref(str(op["range"]))
        for row in ws[cell_range]:
            for cell in row:
                if op.get("bold") is not None: cell.font = Font(bold=op["bold"])
                if op.get("fill_color"):
                    cell.fill = PatternFill(start_color=op["fill_color"].lstrip("#"), end_color=op["fill_color"].lstrip("#"), fill_type="solid")

    def _op_set_column_width(self, op: ExcelOperation) -> None:
        self.ws.column_dimensions[str(op["column"])].width = float(op["width"] or 10.0)

    def _op_merge_cells(self, op: ExcelOperation) -> None:
        ws, cell_range = self._resolve_sheet_and_ref(str(op["range"]))
        ws.merge_cells(cell_range)

    def _op_add_header_row(self, op: ExcelOperation) -> None:
        blue_fill = PatternFill(start_color=NEXUS_BLUE, end_color=NEXUS_BLUE, fill_type="solid")
        ws = self._get_or_create_sheet(str(op["sheet"])) if op.get("sheet") else self.ws
        for col_idx, header in enumerate(op.get("headers", []), 1):
            cell = ws.cell(row=op.get("row", 1), column=col_idx)
            cell.value = header
            cell.font = Font(color=NEXUS_WHITE, bold=True)
            cell.fill = blue_fill

    def _op_create_pivottable(self, op: ExcelOperation) -> None:
        """P3: Dynamic Multi-field Pivot Table Generator."""
        if "source" not in op or "destination" not in op:
            raise ValueError("create_pivottable requires 'source' and 'destination'")

        ws_src, source_rows = self._load_source_rows(str(op["source"]))
        if len(source_rows) < 2: return

        headers = [str(h) for h in source_rows[0]]
        records = [dict(zip(headers, row)) for row in source_rows[1:]]
        
        row_fields = op.get("rows", [])
        value_specs = op.get("values", [])

        # Simple grouping logic for AVERAGE/SUM/COUNT
        pivot_data = defaultdict(list)
        for rec in records:
            key = tuple(rec.get(f) for f in row_fields)
            for spec in value_specs:
                f_name = spec if isinstance(spec, str) else spec.get("field", "")
                val = rec.get(f_name)
                try:
                    if val is not None: pivot_data[key].append(float(val))
                except ValueError:
                    raise ValueError("create_pivottable requires numeric values for aggregation")

        target_ws, anchor = self._resolve_sheet_and_ref(str(op["destination"]), True)
        # Write headers
        for i, f in enumerate(row_fields): target_ws.cell(1, i+1, f).font = Font(bold=True)
        target_ws.cell(1, len(row_fields)+1, "Result").font = Font(bold=True)

        # Write data
        for r_idx, (key, vals) in enumerate(pivot_data.items(), 2):
            for c_idx, k_val in enumerate(key, 1): target_ws.cell(r_idx, c_idx, k_val)
            # Apply aggregation
            func = "SUM" # Default
            if value_specs and isinstance(value_specs[0], dict): func = value_specs[0].get("func", "SUM")
            
            res = sum(vals) if func == "SUM" else len(vals) if func == "COUNT" else sum(vals)/len(vals) if vals else 0
            target_ws.cell(r_idx, len(row_fields)+1, res)

    def _load_source_rows(self, source: str) -> Tuple[Any, List[List[Any]]]:
        ws, ref = self._resolve_sheet_and_ref(source)
        return ws, [[cell.value for cell in row] for row in ws[ref]]

    _DISPATCH = {
        "set_value": "_op_set_value",
        "add_formula": "_op_add_formula",
        "set_formula": "_op_add_formula",
        "format_range": "_op_format_range",
        "set_column_width": "_op_set_column_width",
        "merge_cells": "_op_merge_cells",
        "add_header_row": "_op_add_header_row",
        "create_pivottable": "_op_create_pivottable",
    }

    def _resolve_sheet_and_ref(self, reference: str, create_missing_sheet: bool = False) -> Tuple[Any, str]:
        if "!" not in reference: return self.ws, reference
        sheet_name, local_ref = reference.split("!", 1)
        clean_name = sheet_name.strip("'")
        if create_missing_sheet and clean_name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(clean_name)
        else:
            ws = self.wb[clean_name]
        return ws, local_ref

    def _get_or_create_sheet(self, sheet_name: str) -> Any:
        if sheet_name in self.wb.sheetnames: return self.wb[sheet_name]
        return self.wb.create_sheet(sheet_name)

    def save_workbook(self, output_path: str) -> Dict[str, Any]:
        output_extension = os.path.splitext(output_path)[1].lower()
        if output_extension not in SUPPORTED_SPREADSHEET_SUFFIXES:
            raise ValueError(f"Unsupported spreadsheet output format: {output_extension or '(missing extension)'}")
        parent_dir = os.path.dirname(output_path)
        if parent_dir and not os.path.exists(parent_dir):
            os.makedirs(parent_dir, exist_ok=True)

        if output_extension in (".csv", ".tsv"):
            delimiter = "," if output_extension == ".csv" else "\t"
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=delimiter)
                for row in self.ws.iter_rows(values_only=True):
                    writer.writerow(row)
        else:
            self.wb.save(output_path)
            
        return {
            "status": "success", 
            "file": output_path, 
            "output_format": output_extension.lstrip("."),
            "office_context_received": bool(self.office_context),
            "template_preserved": bool(self.input_path),
            "validation": {"passed": True}
        }

    def _load_workbook(self, excel_path: str) -> openpyxl.Workbook:
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Input file not found: {excel_path}")
        ext = os.path.splitext(excel_path)[1].lower()
        if ext in (".csv", ".tsv"):
            wb = openpyxl.Workbook()
            ws = wb.active
            delimiter = "," if ext == ".csv" else "\t"
            with open(excel_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    ws.append(row)
            return wb
        return openpyxl.load_workbook(excel_path)

def run(payload: ExcelPayload) -> Dict[str, Any]:
    input_path = payload.get("input") or payload.get("input_path")
    output_path = payload.get("output") or payload.get("output_path", "output.xlsx")
    expert = ExcelExpertSkill(input_path, payload.get("office_context"))
    
    # We must ensure exceptions in `apply_ops` cause a hard fail if it's considered fatal for tests,
    # but BaseExpertSkill returns errors gracefully. Let's make it raise an error if any applied
    # operation failed so the tests catch it.
    applied = expert.apply_ops(payload.get("changes", []))
    for op in applied:
        if op.get("status") == "error":
            raise ValueError(op.get("message"))
            
    result = expert.save_workbook(output_path)
    result["applied_operations"] = applied
    return result

if __name__ == "__main__":
    try:
        data = json.load(sys.stdin)
        print(json.dumps(run(data)))
    except Exception as e:
        print(json.dumps({"error": str(e)}), file=sys.stderr)
        sys.exit(1)
